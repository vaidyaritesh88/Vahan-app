"""Parser for primary sales (OEM dispatches) Excel data.

Reads Volume_4W and Volume_2W sheets from the simplified auto database Excel.
Stores model-level monthly data in the primary_sales table.
"""
import logging
from datetime import datetime
from database.schema import get_connection

logger = logging.getLogger(__name__)

# 4W: Map model name prefix to canonical OEM name
PV_OEM_MAP = {
    "Maruti": "Maruti Suzuki", "Marui": "Maruti Suzuki",
    "Tata": "Tata Motors", "Hyundai": "Hyundai",
    "Mahindra": "Mahindra", "Toyota": "Toyota",
    "Honda": "Honda Cars", "Kia": "Kia",
    "MG": "MG Motor", "Skoda": "Skoda VW", "VW": "Skoda VW", "Volkswagen": "Skoda VW",
    "Renault": "Renault", "Nissan": "Nissan",
    "Ford": "Ford", "Datsun": "Nissan", "Fiat": "Fiat",
    "FCA": "Stellantis", "Citroen": "Stellantis",
    "Basalt": "Stellantis", "EC3": "Stellantis",
    "Vans": "Maruti Suzuki",  # Maruti Eeco is the only van
}

# 2W: Map OEM column value to canonical name
TW_OEM_MAP = {
    "Hero": "Hero MotoCorp", "Honda": "Honda 2W",
    "Bajaj": "Bajaj Auto", "TVS": "TVS Motor",
    "Yamaha": "Yamaha", "Suzuki": "Suzuki 2W",
    "Royal Enfield": "Royal Enfield",
    "Ola Electric": "Ola Electric", "Ather": "Ather Energy",
    "Hero Electric": "Hero Electric", "Okinawa": "Okinawa",
    "Ampere": "Ampere", "KTM": "KTM", "Kawasaki": "Kawasaki",
    "HD": "Harley-Davidson", "Triumph": "Triumph",
    "Jawa": "Jawa", "Piaggio": "Piaggio", "UM Lohia": "UM Lohia",
}

# 4W segment headers (used to detect segment boundaries)
PV_SEGMENTS = {
    "Entry Hatchback", "Compact Hatchback", "Premium Hatchback",
    "Super Premium Hatchback", "Compact Sedan", "Upper Sedan",
    "Sub-compact SUV", "Compact SUV", "Mid-SUV", "Premium SUV", "MUV", "Vans",
}

# 2W type headers
TW_TYPES = {"Motorcycle", "Scooter", "Moped", "EV"}

# 2W motorcycle segment headers
TW_MC_SEGMENTS = {
    "Economy Segment", "Entry Executive", "Executive", "Premium",
    "Sports", "Sports Super Premium", "Classic Premium", "Classic Super Premium",
}


def _extract_pv_oem(model_name):
    """Extract OEM from 4W model name (first word)."""
    if not model_name:
        return None, model_name
    prefix = model_name.split(" ", 1)[0]
    oem = PV_OEM_MAP.get(prefix)
    if oem:
        return oem, model_name
    # Try longer prefix for multi-word OEMs
    if model_name.startswith("Royal Enfield"):
        return "Royal Enfield", model_name
    return None, model_name


def parse_volume_4w(filepath):
    """Parse the Volume_4W sheet for PV primary sales data.

    Returns list of dicts: {category, segment, oem_name, model_name, year, month, volume}
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Volume_4W"]

    # Extract dates from row 1 (col B onwards)
    dates = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for v in row[1:]:
            if v is not None and isinstance(v, datetime):
                dates.append(v)

    if not dates:
        logger.error("No dates found in Volume_4W row 1")
        return []

    logger.info(f"Volume_4W: {len(dates)} months from {dates[0]:%b %Y} to {dates[-1]:%b %Y}")

    records = []
    current_segment = "Unknown"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        label = str(row[0]).strip() if row[0] else ""
        if not label or label == "None":
            continue

        # Skip total/summary rows
        if label.startswith("Total "):
            continue

        # Check if this is a segment header
        if label in PV_SEGMENTS:
            current_segment = label
            # Special case: "Vans" is both a segment header AND a data row
            # (Maruti Eeco is the only van). Check if row has numeric data.
            has_numeric = any(
                isinstance(row[c], (int, float)) and row[c] > 0
                for c in range(1, min(len(row), len(dates) + 1))
                if row[c] is not None
            )
            if not has_numeric:
                continue
            # Fall through to treat as data row (OEM = "Maruti" for Vans)

        # Skip "Others" as a standalone row (segment-level other)
        if label == "Others":
            continue

        # Extract OEM from model name
        oem, model = _extract_pv_oem(label)
        if not oem:
            logger.debug(f"R{row_idx}: Could not map OEM for '{label}', skipping")
            continue

        # Extract monthly volumes
        for col_idx, date in enumerate(dates):
            val = row[col_idx + 1] if col_idx + 1 < len(row) else None
            if val is not None and isinstance(val, (int, float)) and val > 0:
                records.append({
                    "category": "PV",
                    "segment": current_segment,
                    "oem_name": oem,
                    "model_name": model,
                    "year": date.year,
                    "month": date.month,
                    "volume": float(val),
                })

    wb.close()
    logger.info(f"Volume_4W: parsed {len(records)} records")
    return records


def parse_volume_2w(filepath):
    """Parse the Volume_2W sheet for 2W primary sales data.

    Returns list of dicts: {category, segment, oem_name, model_name, year, month, volume}
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Volume_2W"]

    # Extract dates from row 2 (col C onwards)
    dates = []
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for v in row[2:]:  # Skip cols A (OEM) and B (Model)
            if v is not None and isinstance(v, datetime):
                dates.append(v)

    if not dates:
        logger.error("No dates found in Volume_2W row 2")
        return []

    logger.info(f"Volume_2W: {len(dates)} months from {dates[0]:%b %Y} to {dates[-1]:%b %Y}")

    records = []
    current_type = "Unknown"
    current_segment = "Unknown"

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), 3):
        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not col_a or col_a == "None":
            continue

        # Skip total rows
        if "Total" in col_a or "Total" in col_b:
            continue

        # Check type headers
        if col_a in TW_TYPES and not col_b:
            current_type = col_a
            if current_type == "Scooter":
                current_segment = "Scooter"
            elif current_type == "Moped":
                current_segment = "Moped"
            elif current_type == "EV":
                current_segment = "EV"
            continue

        # Check motorcycle segment headers
        if col_a in TW_MC_SEGMENTS and not col_b:
            current_segment = col_a
            continue

        # Skip non-OEM rows
        if col_a in TW_MC_SEGMENTS or col_a in TW_TYPES:
            if not col_b:
                continue

        # This is an OEM + Model row
        oem_raw = col_a
        model = col_b if col_b and col_b != "None" else oem_raw

        oem = TW_OEM_MAP.get(oem_raw, oem_raw)

        # Build full model name
        model_name = f"{oem_raw}: {model}" if model != oem_raw else oem_raw

        # Extract monthly volumes
        for col_idx, date in enumerate(dates):
            val = row[col_idx + 2] if col_idx + 2 < len(row) else None
            if val is not None and isinstance(val, (int, float)) and val > 0:
                records.append({
                    "category": "2W",
                    "segment": current_segment,
                    "oem_name": oem,
                    "model_name": model_name,
                    "year": date.year,
                    "month": date.month,
                    "volume": float(val),
                })

    wb.close()
    logger.info(f"Volume_2W: parsed {len(records)} records")
    return records


def load_primary_sales(filepath):
    """Parse Excel and load into primary_sales table (delete + reinsert).

    Returns dict with stats: {pv_records, tw_records, total, date_range}
    """
    logger.info(f"Loading primary sales from: {filepath}")

    # Parse both sheets
    pv_records = parse_volume_4w(filepath)
    tw_records = parse_volume_2w(filepath)
    all_records = pv_records + tw_records

    if not all_records:
        return {"pv_records": 0, "tw_records": 0, "total": 0, "date_range": "N/A"}

    # Delete all existing data and reinsert
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM primary_sales")
    logger.info("Cleared existing primary_sales data")

    now = datetime.now().isoformat()
    batch = []
    for rec in all_records:
        batch.append((
            rec["category"], rec["segment"], rec["oem_name"], rec["model_name"],
            rec["year"], rec["month"], rec["volume"], now,
        ))

    cursor.executemany("""
        INSERT OR REPLACE INTO primary_sales
            (category, segment, oem_name, model_name, year, month, volume, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, batch)

    conn.commit()
    conn.close()

    # Compute stats
    years = sorted(set(r["year"] for r in all_records))
    date_range = f"{min(years)}-{max(years)}"

    stats = {
        "pv_records": len(pv_records),
        "tw_records": len(tw_records),
        "total": len(all_records),
        "date_range": date_range,
    }
    logger.info(f"Loaded {stats['total']:,} primary sales records ({stats['date_range']})")
    return stats
