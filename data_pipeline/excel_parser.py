"""Parse the Vahan Excel tracker file and load data into the database."""
import os
from datetime import datetime
from openpyxl import load_workbook
from config.settings import SHEET_TO_CATEGORY, WEEKLY_SHEET_TO_CATEGORY
from config.oem_normalization import normalize_oem
from database.schema import get_connection


def parse_and_load_excel(file_path):
    """Parse the Excel file and load all data into the database."""
    wb = load_workbook(file_path, data_only=True)
    conn = get_connection()

    stats = {"national": 0, "weekly": 0, "errors": []}

    # Parse category data sheets
    for sheet_name, cat_code in SHEET_TO_CATEGORY.items():
        if sheet_name not in wb.sheetnames:
            stats["errors"].append(f"Sheet '{sheet_name}' not found")
            continue
        ws = wb[sheet_name]
        if cat_code == "CV":
            records = _parse_cv_sheet(ws)
        else:
            records = _parse_category_sheet(ws, cat_code)

        _upsert_national_records(conn, records)
        stats["national"] += len(records)

    # Parse weekly trend sheets
    for sheet_name, cat_code in WEEKLY_SHEET_TO_CATEGORY.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        weekly_records = _parse_weekly_sheet(ws, cat_code)
        _upsert_weekly_records(conn, weekly_records)
        stats["weekly"] += len(weekly_records)

    # Log the load
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO load_log (filename, load_type, status, records_loaded) VALUES (?, ?, ?, ?)",
        (os.path.basename(file_path), "excel_full", "success", stats["national"] + stats["weekly"]),
    )
    conn.commit()
    conn.close()
    return stats


def _get_dates_from_row(ws, row_idx):
    """Extract (column_index, year, month) tuples from the date header row."""
    dates = []
    for col_idx in range(3, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(val, datetime):
            dates.append((col_idx, val.year, val.month))
    return dates


def _parse_category_sheet(ws, category_code):
    """Parse a standard category sheet (PV, 2W, 3W, etc.)."""
    # Row 3 has dates
    dates = _get_dates_from_row(ws, 3)
    if not dates:
        return []

    records = []
    # Read rows starting from row 4 until we hit TOTAL
    for row_idx in range(4, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=2).value
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()

        # Stop at TOTAL row - we only want individual OEM rows + Others
        if raw_name.upper() == "TOTAL":
            break

        oem = normalize_oem(raw_name, category_code)
        if oem is None:
            continue

        for col_idx, year, month in dates:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and isinstance(val, (int, float)) and val > 0:
                records.append({
                    "category_code": category_code,
                    "oem_name": oem,
                    "year": year,
                    "month": month,
                    "volume": float(val),
                })

    # Aggregate: sum volumes for merged OEMs (same normalized name)
    return _aggregate_records(records)


def _parse_cv_sheet(ws):
    """Parse the special CV sheet which has LCV and MHCV sub-sections."""
    dates = _get_dates_from_row(ws, 3)
    if not dates:
        return []

    all_records = []
    current_section = None
    total_count = 0

    for row_idx in range(4, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=2).value
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()

        # Detect section headers
        if raw_name.upper() in ("LCV", "LIGHT COMMERCIAL VEHICLES"):
            current_section = "LCV"
            continue
        if raw_name.upper() in ("MHCV", "MEDIUM & HEAVY COMMERCIAL VEHICLES",
                                 "MEDIUM AND HEAVY COMMERCIAL VEHICLES"):
            current_section = "MHCV"
            continue

        # If we haven't detected a section yet, first rows are likely LCV
        if current_section is None:
            # Check if row 4 has data - if so, first section is LCV
            first_val = ws.cell(row=row_idx, column=3).value
            if isinstance(first_val, (int, float)):
                current_section = "LCV"
            else:
                continue

        if raw_name.upper() == "TOTAL":
            total_count += 1
            if total_count == 1:
                # First TOTAL = end of LCV section
                current_section = None  # reset, next section detection needed
            elif total_count >= 2:
                # Second TOTAL = end of MHCV section, stop parsing raw rows
                break
            continue

        if raw_name.upper() in ("TOTAL COMMERCIAL VEHICLES", "TOTAL CV"):
            break

        if current_section is None:
            # Try to detect: after first TOTAL, look for MHCV header
            continue

        oem = normalize_oem(raw_name, current_section)
        if oem is None:
            continue

        for col_idx, year, month in dates:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and isinstance(val, (int, float)) and val > 0:
                all_records.append({
                    "category_code": current_section,
                    "oem_name": oem,
                    "year": year,
                    "month": month,
                    "volume": float(val),
                })

    # Also create CV (combined) records by summing LCV + MHCV
    aggregated = _aggregate_records(all_records)

    # Build CV totals from LCV + MHCV
    cv_totals = {}
    for r in aggregated:
        key = (r["oem_name"], r["year"], r["month"])
        if key not in cv_totals:
            cv_totals[key] = 0.0
        cv_totals[key] += r["volume"]

    for (oem, year, month), vol in cv_totals.items():
        aggregated.append({
            "category_code": "CV",
            "oem_name": oem,
            "year": year,
            "month": month,
            "volume": vol,
        })

    return aggregated


def _parse_weekly_sheet(ws, category_code):
    """Parse a weekly trends sheet."""
    # Row 5 has weekly dates
    dates = _get_dates_from_row(ws, 5)
    if not dates:
        return []

    records = []
    in_cumulative_section = False
    in_period_section = False
    period_row_for_oem = {}  # oem -> period volumes
    num_days_row = None

    for row_idx in range(6, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=2).value
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()

        # Detect sections
        if raw_name.lower() in ("period volumes", "period volume"):
            in_cumulative_section = False
            in_period_section = True
            continue
        if raw_name.lower() in ("overall ytd market shares", "overall ytd mkt shares",
                                 "ytd market shares", "overall ytd market share"):
            in_period_section = False
            break  # Stop parsing - we don't need market share rows
        if raw_name.lower() in ("no. of days", "no of days", "number of days"):
            num_days_row = row_idx
            continue

        if raw_name.upper() in ("TOTAL", "OTHERS", "GRAND TOTAL"):
            continue

        oem = normalize_oem(raw_name, category_code)
        if oem is None or oem == "Others":
            continue

        if not in_period_section:
            # We're in the cumulative YTD section (first section)
            for col_idx, year, month in dates:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    date_val = ws.cell(row=5, column=col_idx).value
                    if isinstance(date_val, datetime):
                        records.append({
                            "category_code": category_code,
                            "oem_name": oem,
                            "week_ending": date_val.strftime("%Y-%m-%d"),
                            "cumulative_volume": float(val),
                            "period_volume": None,
                            "num_days": None,
                        })

    return records


def _aggregate_records(records):
    """Aggregate records by (category, oem, year, month) - sums volumes for merged OEMs."""
    agg = {}
    for r in records:
        key = (r["category_code"], r["oem_name"], r["year"], r["month"])
        if key not in agg:
            agg[key] = 0.0
        agg[key] += r["volume"]

    return [
        {"category_code": k[0], "oem_name": k[1], "year": k[2], "month": k[3], "volume": v}
        for k, v in agg.items()
    ]


def _upsert_national_records(conn, records):
    """Insert or update national monthly records.

    Before inserting, deletes ALL existing excel-sourced records for the same
    (category_code, year, month) combinations found in the new data.
    This prevents stale OEM entries from persisting when an OEM is removed
    from a sheet between Excel file versions.
    """
    if not records:
        return

    cursor = conn.cursor()

    # Find unique (category_code, year, month) combos in new data
    periods = set()
    for r in records:
        periods.add((r["category_code"], r["year"], r["month"]))

    # Delete existing excel-sourced data for these periods
    for cat, year, month in periods:
        cursor.execute("""
            DELETE FROM national_monthly
            WHERE category_code = ? AND year = ? AND month = ? AND source = 'excel'
        """, (cat, year, month))

    # Insert fresh data
    for r in records:
        cursor.execute("""
            INSERT INTO national_monthly (category_code, oem_name, year, month, volume, source)
            VALUES (?, ?, ?, ?, ?, 'excel')
            ON CONFLICT(category_code, oem_name, year, month)
            DO UPDATE SET volume=excluded.volume, updated_at=CURRENT_TIMESTAMP
        """, (r["category_code"], r["oem_name"], r["year"], r["month"], r["volume"]))
    conn.commit()


def _upsert_weekly_records(conn, records):
    """Insert or update weekly trend records."""
    cursor = conn.cursor()
    for r in records:
        cursor.execute("""
            INSERT INTO weekly_trends (category_code, oem_name, week_ending, cumulative_volume, period_volume, num_days)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(category_code, oem_name, week_ending)
            DO UPDATE SET cumulative_volume=excluded.cumulative_volume, period_volume=excluded.period_volume, num_days=excluded.num_days
        """, (r["category_code"], r["oem_name"], r["week_ending"],
              r["cumulative_volume"], r.get("period_volume"), r.get("num_days")))
    conn.commit()
